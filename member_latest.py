from operator import concat
from pprint import pprint
import requests
import lxml.html
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
import os.path
from datetime import datetime
from Crypto.Cipher import AES
from base64 import b64decode
import re

zip_lang = 'zip_language.xlsx'
df = pd.read_excel(zip_lang)
base_url = 'https://www.sia.ch/'
page = input('Enter Page number to Start: ')
n = input('Enter Row number to Start: ')
file_exists = os.path.exists('member.xlsx')

page = int(page)-1
n = int(n)-1
if file_exists:
    deleteData = input('Want to delete old Data (y/n): ')
    if (str(deleteData))=='y':
        wb = load_workbook(filename = "member.xlsx")
        ws = wb["member"]
        mo_ws = wb['member_office']
        print('File found deleting old data')
        ws.delete_rows(2,ws.max_row+1)
        mo_ws.delete_rows(2,mo_ws.max_row+1)
        wb.save("member.xlsx")
    fe_flag = 1
else:
    fe_flag = 0

def pad(data, ks):
    pad_len = (ks - (len(data) % ks)) % ks 
    return data + (b'\x00' * pad_len)

def kdf(pwd, keySize): 
    if keySize != 16 and keySize != 24 and keySize != 32:
        raise ValueError("Wrong keysize") 
    keyPadded = pwd[:keySize] if len(pwd) >= keySize else pad(pwd, keySize)
    aes = AES.new(key=keyPadded, mode=AES.MODE_ECB) 
    key = aes.encrypt(keyPadded[:16])
    if keySize > 16:
        key = (key + key)[:keySize]
    return key

def get_indv(page,df,fe_flag,n):
    ids = n
    row = n+2
    url = base_url +'fr/affiliation/liste-des-membres/membres-individuels/nc/1/?tx_updsiafeuseradmin_pi1%5BdisplaySearchResult%5D=1&tx_updsiafeuseradmin_pi1%5Bpointer%5D='
    while url:
        page_url = url + str(page)
        html = requests.get(page_url)
        while n<50:
            doc = lxml.html.fromstring(html.content)
            indv_zip = doc.xpath('//table//tr['+str(n+2)+']/td[4]/text()') #extracted zip list with 1 element
            if indv_zip: #check for empty zip field
                indv_zip=indv_zip[0] #zip_code
                if (df['ZIP_CODE'].eq(int(indv_zip))).any():
                    indv_lang = df.loc[df['ZIP_CODE'] == int(indv_zip)].LANGUAGE.item() #get lang by comparing zip with excel
                else:
                    indv_lang = 'FR'
                indv_mem_url = doc.xpath('//table//a/@href')[n]
                indv_mem_url_lang = base_url + indv_mem_url.replace("/fr/", str(indv_lang).lower() + '/' ) #member url with language

                indv_html = requests.get(indv_mem_url_lang) #request by mem_url
                indv_doc = lxml.html.fromstring(indv_html.content)

                indv_full_address = indv_doc.xpath('//table//tr[2]/td/text()') #get full address block 
                indv_full_address_clean = clean_list(indv_full_address)
                while(len(indv_full_address_clean)<5):
                        indv_full_address_clean.append('')
                join_indv_full_address_clean = " ".join(indv_full_address_clean)
                
                #***implement later
                data_contact = indv_doc.xpath('//@data-contact')[0]
                data_secret = indv_doc.xpath('//@data-secr')[0]
                #print(data_contact[0])
                #print(data_secret[0])

                ciphertext = b64decode(data_contact)
                nc = ciphertext[:8]
                data = ciphertext[8:]

                keySize = 32
                pwd = data_secret #from data-secr
                key = kdf(pwd.encode('utf-8'), keySize) 
                aes = AES.new(key=key, mode=AES.MODE_CTR, nonce=nc) 
                res = aes.decrypt(data)
                result = res.decode('utf-8')
                tel = result[0:13]
                email = re.findall(r'[\w.+-]+@[\w-]+\.[\w.-]+', result)
                if email:
                    email = email[0]
                else:
                    email=''
                fax=result[44:57]
                website = re.search('_blank">(.*)</a><br />', result)
                if website:
                    website = website.group(1)
                else:
                    website=''
                #print(result, tel)

                contact=[email,tel,fax,website]

                job = indv_doc.xpath('//table//tr[6]/td[2]/text()')
                if job:
                    job = clean_list(job)
                else:
                    job=['']
                sector = indv_doc.xpath('//table//tr[7]/td[2]/text()')
                if sector:
                    sector = clean_list(sector)
                else:
                    sector=['']
                
                group = indv_doc.xpath('//table//tr[8]/td[2]/text()')
                if group:
                    group = clean_list(group)
                else:
                    group=['']
                
                section = indv_doc.xpath('//table//tr[9]/td[2]/text()')
                if section:
                    section = clean_list(section)
                else:
                    section=['']

                #print(join_indv_full_address_clean, contact)
                #print(job, sector, group, section)
                wdf = pd.DataFrame([[ids+1, indv_mem_url_lang, indv_lang, join_indv_full_address_clean,indv_full_address_clean[0], indv_full_address_clean[1], indv_full_address_clean[2], indv_full_address_clean[3], indv_full_address_clean[4], indv_zip, contact[0], contact[1], contact[2], contact[2], job[0], sector[0], group[0], section[0]]], columns=["ID","URL", "LANGUGE", "FULL_ADDRESS", "GENDER", "NAME", "EDUCATION", "ADDRESS", "CITY", "ZIP_CODE", "EMAIL", "TEL", "FAX", "WEBSITE", "JOB", "SECTOR", "GROUP", "SECTION"])

                now = datetime.now()
                dts = now.strftime("%d/%m/%Y %H:%M")
                mo_wdf = pd.DataFrame([[ids+1, ids+1, '', dts]], columns=["ID","MEMBER_ID", "OFFICE_ID", "COLLECTED_AT"])

                office_wdf = pd.DataFrame(columns=["ID","URL", "LANGUGE", "FULL_ADDRESS", "NAME", "ADDRESS", "CITY", "ZIP_CODE", "EMAIL", "TEL", "FAX", "WEBSITE", "SECTOR"])
                #with pd.ExcelWriter("member.xlsx") as writer:
                    #wdf.to_excel(writer, sheet_name='member', index=False)
                #print(wdf)
                if fe_flag==1:
                    wb = load_workbook(filename = "member.xlsx")
                    ws = wb["member"]
                    mows = wb["member_office"]
                    for r in dataframe_to_rows(wdf, index=False, header=False):
                        ws.append(r)
                    mows.cell(row=int(row), column=1).value = ids+1
                    mows.cell(row=int(row), column=2).value = ids+1
                    mows.cell(row=int(row), column=4).value = dts
                    #for mo in dataframe_to_rows(mo_wdf, index=False, header=False):
                        #mows.append(mo)

                    wb.save("member.xlsx")
                    wb.close
                    print("Saving info of page: " + str(page+1) +"  member: " + str(n+1) +" in excel")
                else:
                    with pd.ExcelWriter("member.xlsx") as writer:
                        wdf.to_excel(writer, sheet_name='member', index=False)
                        mo_wdf.to_excel(writer, sheet_name='member_office', index=False)
                        office_wdf.to_excel(writer, sheet_name='office', index=False)
                    print('Creating New Excel')
                    fe_flag=1
                    print("Saving info of page: " + str(page+1) +"  member: " + str(n+1) +" in excel")
                row+=1
                
            else: #if zip not found
                indv_lang = 'FR'
                indv_mem_url = doc.xpath('//table//a/@href')[n]
                indv_mem_url_lang = base_url + indv_mem_url.replace("/fr/", str(indv_lang).lower() + '/' ) #member url with language

                indv_html = requests.get(indv_mem_url_lang) #request by mem_url
                indv_doc = lxml.html.fromstring(indv_html.content)

                indv_full_address = indv_doc.xpath('//table//tr[2]/td/text()') #get full address block 
                indv_full_address_clean = clean_list(indv_full_address)
                while(len(indv_full_address_clean)<5):
                        indv_full_address_clean.append('')
                join_indv_full_address_clean = " ".join(indv_full_address_clean)

                #***implement later
                data_contact = indv_doc.xpath('//@data-contact')[0]
                data_secret = indv_doc.xpath('//@data-secr')[0]
                #print(data_contact[0])
                #print(data_secret[0])

                ciphertext = b64decode(data_contact)
                nc = ciphertext[:8]
                data = ciphertext[8:]

                keySize = 32
                pwd = data_secret #from data-secr
                key = kdf(pwd.encode('utf-8'), keySize) 
                aes = AES.new(key=key, mode=AES.MODE_CTR, nonce=nc) 
                res = aes.decrypt(data)
                result = res.decode('utf-8')
                tel = result[0:13]
                email = re.findall(r'[\w.+-]+@[\w-]+\.[\w.-]+', result)
                if email:
                    email = email[0]
                else:
                    email=''
                fax=result[44:57]
                website = re.search('_blank">(.*)</a><br />', result)
                if website:
                    website = website.group(1)
                else:
                    website=''
                #print(result, tel)

                contact=[email,tel,fax,website]


                job = indv_doc.xpath('//table//tr[6]/td[2]/text()')
                if job:
                    job = clean_list(job)
                else:
                    job=['']

                sector = indv_doc.xpath('//table//tr[7]/td[2]/text()')
                if sector:
                    sector = clean_list(sector)
                else:
                    sector=['']
                
                group = indv_doc.xpath('//table//tr[8]/td[2]/text()')
                if group:
                    group = clean_list(group)
                else:
                    group=['']
                
                section = indv_doc.xpath('//table//tr[9]/td[2]/text()')
                if section:
                    section = clean_list(section)
                else:
                    section=['']

                #print(join_indv_full_address_clean, contact)
                #print(job, sector, group, section)
                wdf = pd.DataFrame([[ids+1, indv_mem_url_lang, indv_lang, join_indv_full_address_clean,indv_full_address_clean[0], indv_full_address_clean[1], indv_full_address_clean[2], indv_full_address_clean[3], indv_full_address_clean[4], '', contact[0], contact[1], contact[2], contact[3], job[0], sector[0], group[0], section[0]]], columns=["ID","URL", "LANGUGE", "FULL_ADDRESS", "GENDER", "NAME", "EDUCATION", "ADDRESS", "CITY", "ZIP_CODE", "EMAIL", "TEL", "FAX", "WEBSITE", "JOB", "SECTOR", "GROUP", "SECTION"])

                now = datetime.now()
                dts = now.strftime("%d/%m/%Y %H:%M")
                mo_wdf = pd.DataFrame(columns=["ID","MEMBER_ID", "OFFICE_ID", "COLLECTED_AT"])

                office_wdf = pd.DataFrame(columns=["ID","URL", "LANGUGE", "FULL_ADDRESS", "NAME", "ADDRESS", "CITY", "ZIP_CODE", "EMAIL", "TEL", "FAX", "WEBSITE", "SECTOR"])
                #with pd.ExcelWriter("member.xlsx") as writer:
                    #wdf.to_excel(writer, sheet_name='member', index=False)
                #print(wdf)
                if fe_flag==1:
                    wb = load_workbook(filename = "member.xlsx")
                    ws = wb["member"]
                    mows = wb["member_office"]

                    for r in dataframe_to_rows(wdf, index=False, header=False):
                        ws.append(r)
                    mows.cell(row=int(row), column=1).value = ids+1
                    mows.cell(row=int(row), column=2).value = ids+1
                    mows.cell(row=int(row), column=4).value = dts
                    #for mo in dataframe_to_rows(mo_wdf, index=False, header=False):
                        #mows.append(mo)

                    wb.save("member.xlsx")
                    wb.close
                    print("Saving info of page: " + str(page+1) +"  member: " + str(n+1) +" in excel")
                else:
                    with pd.ExcelWriter("member.xlsx") as writer:
                        wdf.to_excel(writer, sheet_name='member', index=False)
                        mo_wdf.to_excel(writer, sheet_name='member_office', index=False)
                        office_wdf.to_excel(writer, sheet_name='office', index=False)
                    print('Creating New Excel')
                    fe_flag=1
                    print("Saving info of page: " + str(page+1) +"  member: " + str(n+1) +" in excel")
                row+=1
            ids+=1
            n+=1
        page=page+1
        n=0
def clean_list(list2clean):
    clean_list = []
    #clean full address
    for element in list2clean:
        clean_list.append(element.strip())
    return list(filter(lambda e: e != '', clean_list))

print(get_indv(page,df,fe_flag,n))