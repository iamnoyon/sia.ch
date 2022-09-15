from pprint import pprint
import requests
import lxml.html
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
import os.path

zip_lang = 'zip_language.xlsx'
df = pd.read_excel(zip_lang)
base_url = 'https://www.sia.ch/'
page=0
n=0
file_exists = os.path.exists('member.xlsx')
if file_exists:
    wb = load_workbook(filename = "member.xlsx")
    if 'office' in wb.sheetnames:
        ows = wb["office"]
        print('file found deleting old data')
        ows.delete_rows(2,ows.max_row+1)
        wb.save("member.xlsx")
        fe_flag = 1
else:
    fe_flag = 0

def get_indv(page,df,fe_flag,n):
    ids = n
    office_url = base_url +'fr/membership/member-directory/corporate-members/nc/1/?tx_updsiafeuseradmin_pi1%5BdisplaySearchResult%5D=1&tx_updsiafeuseradmin_pi1%5Bpointer%5D='
    while office_url:
        office_page_url = office_url + str(page)
        office_html = requests.get(office_page_url)
        while n<50:
            office_doc = lxml.html.fromstring(office_html.content)

            indv_zip = office_doc.xpath('//table//tr['+str(n+2)+']/td[4]/text()') #extracted zip list with 1 element
            if indv_zip: #check for empty zip field
                indv_zip=indv_zip[0] #zip_code
                if (df['ZIP_CODE'].eq(int(indv_zip))).any():
                    indv_lang = df.loc[df['ZIP_CODE'] == int(indv_zip)].LANGUAGE.item() #get lang by comparing zip with excel
                else:
                    indv_lang = 'FR'
                indv_office_url = office_doc.xpath('//table//td[1]//a/@href')[n]
                indv_office_url_lang = base_url + indv_office_url.replace("/fr/", str(indv_lang).lower() + '/' ) #member url with language

                indv_office_html = requests.get(indv_office_url_lang) #request by mem_url

                indv_office_doc = lxml.html.fromstring(indv_office_html.content)

                indv_office_full_address = indv_office_doc.xpath('//table//tr[2]/td/text()') #get full address block 
                indv_office_full_address_clean = clean_list(indv_office_full_address)
                while(len(indv_office_full_address_clean)<4):
                        indv_office_full_address_clean.append('')
                join_indv_office_full_address_clean = " ".join(indv_office_full_address_clean)

                office_sector_list = indv_office_doc.xpath('//tr[6]/td/ul//text()')
                if office_sector_list:
                    office_sector = ",".join(office_sector_list)
                else:
                    office_sector = ''

                #***implement later
                contact = indv_office_doc.xpath('//table//tr[4]/td[2]/div/text()')

                #print(join_indv_full_address_clean, contact)
                #print(job, sector, group, section)
                office_wdf = pd.DataFrame([[ids+1, indv_office_url_lang, indv_lang, join_indv_office_full_address_clean,indv_office_full_address_clean[0], indv_office_full_address_clean[1], indv_office_full_address_clean[2], indv_zip, '','','','',office_sector]], columns=["ID","URL", "LANGUGE", "FULL_ADDRESS", "NAME", "ADDRESS", "CITY", "ZIP_CODE", "EMAIL", "TEL", "FAX", "WEBSITE", "SECTOR"])
                #with pd.ExcelWriter("member.xlsx") as writer:
                    #wdf.to_excel(writer, sheet_name='member', index=False)
                #print(wdf)
                if fe_flag==1:
                    wb = load_workbook(filename = "member.xlsx")
                    ows = wb["office"]
                    for o in dataframe_to_rows(office_wdf, index=False, header=False):
                        ows.append(o)
                    wb.save("member.xlsx")
                    wb.close
                    print("Saving info of page: " + str(page+1) +"  office: " + str(n+1) +" in excel")
                else:
                    with pd.ExcelWriter("member.xlsx") as writer:
                        office_wdf.to_excel(writer, sheet_name='office', index=False)
                    print('Creating New Excel')
                    fe_flag=1
                    print("Saving info of page: " + str(page+1) +"  Office: " + str(n+1) +" in excel")


            else: #if zip not found
                indv_lang = 'FR'
                indv_office_url = office_doc.xpath('//table//td[1]//a/@href')[n]
                indv_office_url_lang = base_url + indv_office_url.replace("/fr/", str(indv_lang).lower() + '/' ) #member url with language


                indv_office_html = requests.get(indv_office_url_lang) #request by mem_url

                indv_office_doc = lxml.html.fromstring(indv_office_html.content)

                indv_office_full_address = indv_office_doc.xpath('//table//tr[2]/td/text()') #get full address block 
                indv_office_full_address_clean = clean_list(indv_office_full_address)
                while(len(indv_office_full_address_clean)<4):
                        indv_office_full_address_clean.append('')
                join_indv_office_full_address_clean = " ".join(indv_office_full_address_clean)

                office_sector_list = indv_office_doc.xpath('//table//tr[6]/td/ul//text()')
                if office_sector_list:
                    office_sector = ",".join(office_sector_list)
                else:
                    office_sector = ''

                #***implement later
                contact = indv_office_doc.xpath('//table//tr[4]/td[2]/div/text()')

                #print(join_indv_full_address_clean, contact)
                #print(job, sector, group, section)
                office_wdf = pd.DataFrame([[ids+1, indv_office_url_lang, indv_lang, join_indv_office_full_address_clean,indv_office_full_address_clean[0], indv_office_full_address_clean[1], indv_office_full_address_clean[2], indv_zip, '','','','',office_sector]], columns=["ID","URL", "LANGUGE", "FULL_ADDRESS", "NAME", "ADDRESS", "CITY", "ZIP_CODE", "EMAIL", "TEL", "FAX", "WEBSITE", "SECTOR"])
                #with pd.ExcelWriter("member.xlsx") as writer:
                    #wdf.to_excel(writer, sheet_name='member', index=False)
                #print(wdf)
                if fe_flag==1:
                    wb = load_workbook(filename = "member.xlsx")
                    ows = wb["office"]
                    for o in dataframe_to_rows(office_wdf, index=False, header=False):
                        ows.append(o)
                    wb.save("member.xlsx")
                    wb.close
                    print("Saving info of page: " + str(page+1) +"  office: " + str(n+1) +" in excel")
                else:
                    with pd.ExcelWriter("member.xlsx") as writer:
                        office_wdf.to_excel(writer, sheet_name='office', index=False)
                    print('Creating New Excel')
                    fe_flag=1
                    print("Saving info of page: " + str(page+1) +"  Office: " + str(n+1) +" in excel")

            ids+=1
            n+=1
        page=page+1
        n=0
def clean_list(list2clean):
    clean_list = []
    #clean full address
    for element in list2clean:
        clean_list.append(element.strip())
    return list(filter(lambda e: e != '', clean_list))

print(get_indv(page,df,fe_flag,n))