from pprint import pprint
import requests
import lxml.html
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
import os.path
from datetime import datetime
from Crypto.Cipher import AES
from base64 import b64decode
import re

zip_lang = 'zip_language.xlsx'
df = pd.read_excel(zip_lang)
base_url = 'https://www.sia.ch/'
page = input('Enter Page number to Start: ')
n = input('Enter Row number to Start: ')
file_exists = os.path.exists('member.xlsx')
page = int(page)-1
n = int(n)-1
ids = n + (50*page)
rowc=n+2

def empty_cell_counter(mo_ws):
    ec_count = 0
    for x in mo_ws['C']:
        if x.value is None:
            ec_count+=1
    return ec_count

if file_exists:
    deleteData = input('Want to delete old Data (y/n): ')
    if (str(deleteData))=='y':
        wb = load_workbook(filename = "member.xlsx")
        ws = wb["member"]
        mo_ws = wb['member_office']
        print('File found deleting old data')
        ws.delete_rows(2,ws.max_row+1)
        mo_ws.delete_rows(2,mo_ws.max_row+1)
        wb.save("member.xlsx")
        fe_flag = 1
    elif (str(deleteData))=='n':
        wb = load_workbook(filename = "member.xlsx")
        mo_ws = wb['member_office']
        temp = empty_cell_counter(mo_ws)
        #print(temp)
        rowc=len(mo_ws['C'])-temp+1
        #for row in mo_ws.iter_rows():
            #for cell in row:
                #if cell.value == int(ids):
                    #rowc=row[0].value+1
                    #break
        fe_flag=1
else:
    fe_flag = 0


def pad(data, ks):
    pad_len = (ks - (len(data) % ks)) % ks 
    return data + (b'\x00' * pad_len)

def kdf(pwd, keySize): 
    if keySize != 16 and keySize != 24 and keySize != 32:
        raise ValueError("Wrong keysize") 
    keyPadded = pwd[:keySize] if len(pwd) >= keySize else pad(pwd, keySize)
    aes = AES.new(key=keyPadded, mode=AES.MODE_ECB) 
    key = aes.encrypt(keyPadded[:16])
    if keySize > 16:
        key = (key + key)[:keySize]
    return key

def get_indv(page,df,fe_flag,n,ids,rowc):
    office_url = base_url +'fr/membership/member-directory/corporate-members/nc/1/?tx_updsiafeuseradmin_pi1%5BdisplaySearchResult%5D=1&tx_updsiafeuseradmin_pi1%5Bpointer%5D='
    while office_url:
        office_page_url = office_url + str(page)
        office_html = requests.get(office_page_url)
        while n<50:
            office_doc = lxml.html.fromstring(office_html.content)

            office_indv_zip = office_doc.xpath('//table//tr['+str(n+2)+']/td[4]/text()') #extracted zip list with 1 element
            if office_indv_zip: #check for empty zip field
                office_indv_zip=office_indv_zip[0] #zip_code
                if (df['ZIP_CODE'].eq(int(office_indv_zip))).any():
                    office_indv_lang = df.loc[df['ZIP_CODE'] == int(office_indv_zip)].LANGUAGE.item() #get lang by comparing zip with excel
                else:
                    office_indv_lang = 'FR'
                indv_office_url = office_doc.xpath('//table//td[1]//a/@href')[n]
                indv_office_url_lang = base_url + indv_office_url.replace("/fr/", str(office_indv_lang).lower() + '/' ) #member url with language

                indv_office_html = requests.get(indv_office_url_lang) #request by mem_url

                indv_office_doc = lxml.html.fromstring(indv_office_html.content)

                indv_office_full_address = indv_office_doc.xpath('//table//tr[2]/td/text()') #get full address block 
                indv_office_full_address_clean = clean_list(indv_office_full_address)
                while(len(indv_office_full_address_clean)<4):
                        indv_office_full_address_clean.append('')
                join_indv_office_full_address_clean = " ".join(indv_office_full_address_clean)

                office_sector_list = indv_office_doc.xpath('//tr[6]/td/ul//text()')
                if office_sector_list:
                    office_sector = ",".join(office_sector_list)
                else:
                    office_sector = ''

                #***implement later
                data_contact = indv_office_doc.xpath('//@data-contact')
                if data_contact:
                    data_contact=data_contact[0]
                else:
                    data_contact=''
                data_secret = indv_office_doc.xpath('//@data-secr')
                if data_secret:
                    data_secret=data_secret[0]
                else:
                    data_secret = ''
                #print(data_contact[0])
                #print(data_secret[0])

                ciphertext = b64decode(data_contact)
                nc = ciphertext[:8]
                data = ciphertext[8:]

                keySize = 32
                pwd = data_secret #from data-secr
                key = kdf(pwd.encode('utf-8'), keySize) 
                aes = AES.new(key=key, mode=AES.MODE_CTR, nonce=nc) 
                res = aes.decrypt(data)
                result = res.decode('utf-8')
                #tel = result[0:13]
                email = re.findall(r'[\w.+-]+@[\w-]+\.[\w.-]+', result)
                if email:
                    email = email[0]
                else:
                    email=''
                phone_fax=re.findall(r"(\+\d\d\s*\d(?:\d)+)",result)
                if phone_fax and len(phone_fax)==2:
                    tel = phone_fax[0]
                    fax = phone_fax[1]
                elif phone_fax and len(phone_fax)==1:
                    tel = phone_fax[0]
                    fax = ''
                else:
                    tel = ''
                    fax = ''
                website = re.search('_blank">(.*)</a><br />', result)
                if website:
                    website = website.group(1)
                else:
                    website=''
                #print(result, tel)

                #contact=[tel,email,fax,website]
                #print(join_indv_full_address_clean, contact)
                #print(job, sector, group, section)
                wdf = pd.DataFrame(columns=["ID","URL", "LANGUGE", "FULL_ADDRESS", "GENDER", "NAME", "EDUCATION", "ADDRESS", "CITY", "ZIP_CODE", "CONTACT", "JOB", "SECTOR", "GROUP", "SECTION"])

                now = datetime.now()
                dts = now.strftime("%d/%m/%Y %H:%M")
                mo_wdf = pd.DataFrame([[ids+1,'',ids+1,dts]], columns=["ID","MEMBER_ID", "OFFICE_ID", "COLLECTED_AT"])

                office_wdf = pd.DataFrame([[ids+1, indv_office_url_lang, office_indv_lang, join_indv_office_full_address_clean,indv_office_full_address_clean[0], indv_office_full_address_clean[1], indv_office_full_address_clean[2], office_indv_zip, email,tel,fax,website,office_sector]], columns=["ID","URL", "LANGUGE", "FULL_ADDRESS", "NAME", "ADDRESS", "CITY", "ZIP_CODE", "EMAIL", "TEL", "FAX", "WEBSITE", "SECTOR"])
                #with pd.ExcelWriter("member.xlsx") as writer:
                    #wdf.to_excel(writer, sheet_name='member', index=False)
                #print(wdf)
                if fe_flag==1:
                    wb = load_workbook(filename = "member.xlsx")
                    ows = wb["office"]
                    mows = wb["member_office"]
                    for o in dataframe_to_rows(office_wdf, index=False, header=False):
                        ows.append(o)
                    mows.cell(row=int(rowc), column=1).value = ids+1
                    mows.cell(row=int(rowc), column=3).value = ids+1
                    mows.cell(row=int(rowc), column=4).value = dts
                    wb.save("member.xlsx")
                    wb.close
                    print("Saving info of page: " + str(page+1) +"  office: " + str(n+1) +" in excel")
                else:
                    with pd.ExcelWriter("member.xlsx") as writer:
                        wdf.to_excel(writer, sheet_name='member', index=False)
                        mo_wdf.to_excel(writer, sheet_name='member_office', index=False)
                        office_wdf.to_excel(writer, sheet_name='office', index=False)
                    print('Creating New Excel')
                    fe_flag=1
                    print("Saving info of page: " + str(page+1) +"  Office: " + str(n+1) +" in excel")
                rowc+=1

            else: #if zip not found
                office_indv_lang = 'FR'
                indv_office_url = office_doc.xpath('//table//td[1]//a/@href')[n]
                indv_office_url_lang = base_url + indv_office_url.replace("/fr/", str(office_indv_lang).lower() + '/' ) #member url with language


                indv_office_html = requests.get(indv_office_url_lang) #request by mem_url

                indv_office_doc = lxml.html.fromstring(indv_office_html.content)

                indv_office_full_address = indv_office_doc.xpath('//table//tr[2]/td/text()') #get full address block 
                indv_office_full_address_clean = clean_list(indv_office_full_address)
                while(len(indv_office_full_address_clean)<4):
                        indv_office_full_address_clean.append('')
                join_indv_office_full_address_clean = " ".join(indv_office_full_address_clean)

                office_sector_list = indv_office_doc.xpath('//table//tr[6]/td/ul//text()')
                if office_sector_list:
                    office_sector = ",".join(office_sector_list)
                else:
                    office_sector = ''

                #***implement later
                data_contact = indv_office_doc.xpath('//@data-contact')
                if data_contact:
                    data_contact=data_contact[0]
                else:
                    data_contact=''
                data_secret = indv_office_doc.xpath('//@data-secr')
                if data_secret:
                    data_secret=data_secret[0]
                else:
                    data_secret = ''
                #print(data_contact[0])
                #print(data_secret[0])

                ciphertext = b64decode(data_contact)
                nc = ciphertext[:8]
                data = ciphertext[8:]

                keySize = 32
                pwd = data_secret #from data-secr
                key = kdf(pwd.encode('utf-8'), keySize) 
                aes = AES.new(key=key, mode=AES.MODE_CTR, nonce=nc) 
                res = aes.decrypt(data)
                result = res.decode('utf-8')
                #tel = result[0:13]
                email = re.findall(r'[\w.+-]+@[\w-]+\.[\w.-]+', result)
                if email:
                    email = email[0]
                else:
                    email=''
                phone_fax=re.findall(r"(\+\d\d\s*\d(?:\d)+)",result)
                if phone_fax and len(phone_fax)==2:
                    tel = phone_fax[0]
                    fax = phone_fax[1]
                elif phone_fax and len(phone_fax)==1:
                    tel = phone_fax[0]
                    fax = ''
                else:
                    tel = ''
                    fax = ''
                website = re.search('_blank">(.*)</a><br />', result)
                if website:
                    website = website.group(1)
                else:
                    website=''
                #print(result, tel)

                #contact=[tel,email,fax,website]

                #print(join_indv_full_address_clean, contact)
                #print(job, sector, group, section)
                wdf = pd.DataFrame(columns=["ID","URL", "LANGUGE", "FULL_ADDRESS", "GENDER", "NAME", "EDUCATION", "ADDRESS", "CITY", "ZIP_CODE", "CONTACT", "JOB", "SECTOR", "GROUP", "SECTION"])

                now = datetime.now()
                dts = now.strftime("%d/%m/%Y %H:%M")
                mo_wdf = pd.DataFrame([[ids+1,'',ids+1,dts]], columns=["ID","MEMBER_ID", "OFFICE_ID", "COLLECTED_AT"])

                office_wdf = pd.DataFrame([[ids+1, indv_office_url_lang, office_indv_lang, join_indv_office_full_address_clean,indv_office_full_address_clean[0], indv_office_full_address_clean[1], indv_office_full_address_clean[2], '', email,tel,fax,website,office_sector]], columns=["ID","URL", "LANGUGE", "FULL_ADDRESS", "NAME", "ADDRESS", "CITY", "ZIP_CODE", "EMAIL", "TEL", "FAX", "WEBSITE", "SECTOR"])
                #with pd.ExcelWriter("member.xlsx") as writer:
                    #wdf.to_excel(writer, sheet_name='member', index=False)
                #print(wdf)
                if fe_flag==1:
                    wb = load_workbook(filename = "member.xlsx")
                    ows = wb["office"]
                    mows = wb["member_office"]
                    for o in dataframe_to_rows(office_wdf, index=False, header=False):
                        ows.append(o)
                    mows.cell(row=int(rowc), column=1).value = ids+1
                    mows.cell(row=int(rowc), column=3).value = ids+1
                    mows.cell(row=int(rowc), column=4).value = dts
                    wb.save("member.xlsx")
                    wb.close
                    print("Saving info of page: " + str(page+1) +"  office: " + str(n+1) +" in excel")
                else:
                    with pd.ExcelWriter("member.xlsx") as writer:
                        wdf.to_excel(writer, sheet_name='member', index=False)
                        mo_wdf.to_excel(writer, sheet_name='member_office', index=False)
                        office_wdf.to_excel(writer, sheet_name='office', index=False)
                    print('Creating New Excel')
                    fe_flag=1
                    print("Saving info of page: " + str(page+1) +"  Office: " + str(n+1) +" in excel")
                rowc+=1
            ids+=1
            n+=1
        page=page+1
        n=0
def clean_list(list2clean):
    clean_list = []
    #clean full address
    for element in list2clean:
        clean_list.append(element.strip())
    return list(filter(lambda e: e != '', clean_list))

print(get_indv(page,df,fe_flag,n,ids,rowc))