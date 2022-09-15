from pprint import pprint
import requests
import lxml.html
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import load_workbook
import os.path
from datetime import datetime

zip_lang = 'zip_language.xlsx'
df = pd.read_excel(zip_lang)
base_url = 'https://www.sia.ch/'
page=0
n=0
file_exists = os.path.exists('member.xlsx')
if file_exists:
    wb = load_workbook(filename = "member.xlsx")
    ws = wb["member"]
    mows = wb["member_office"]
    ows = wb["office"]
    print('file found deleting old data')
    ws.delete_rows(2,ws.max_row+1)
    mows.delete_rows(2,mows.max_row+1)
    ows.delete_rows(2,ows.max_row+1)
    wb.save("member.xlsx")
    fe_flag = 1
else:
    fe_flag = 0

def get_indv(page,df,fe_flag,n):
    ids = n
    url = base_url +'fr/affiliation/liste-des-membres/membres-individuels/nc/1/?tx_updsiafeuseradmin_pi1%5BdisplaySearchResult%5D=1&tx_updsiafeuseradmin_pi1%5Bpointer%5D='
    office_url = base_url +'fr/affiliation/liste-des-membres/membres-bureaux/nc/1/?tx_updsiafeuseradmin_pi1%5BdisplaySearchResult%5D='

    while url and office_url:
        page_url = url + str(page)
        html = requests.get(page_url)

        office_page_url = office_url + str(page)
        office_html = requests.get(office_page_url)
        
        while n<50:
            doc = lxml.html.fromstring(html.content)
            office_doc = lxml.html.fromstring(office_html.content)

            indv_zip = doc.xpath('//table//tr['+str(n+2)+']/td[4]/text()') #extracted zip list with 1 element
            office_indv_zip = office_doc.xpath('//table//tr['+str(n+2)+']/td[4]/text()') #extracted zip list with 1 element

            if indv_zip and office_indv_zip: #check for empty zip field
                indv_zip=indv_zip[0] #zip_code
                office_indv_zip=office_indv_zip[0] #zip_code
                if (df['ZIP_CODE'].eq(int(indv_zip))).any():
                    indv_lang = df.loc[df['ZIP_CODE'] == int(indv_zip)].LANGUAGE.item() #get lang by comparing zip with excel
                else:
                    indv_lang = 'FR'

                if (df['ZIP_CODE'].eq(int(office_indv_zip))).any():
                    office_indv_lang = df.loc[df['ZIP_CODE'] == int(office_indv_zip)].LANGUAGE.item() #get lang by comparing zip with excel
                else:
                    office_indv_lang = 'FR'

                indv_mem_url = doc.xpath('//table//a/@href')[n]
                indv_office_url = doc.xpath('//table//td[1]//a/@href')[n]

                indv_mem_url_lang = base_url + indv_mem_url.replace("/fr/", str(indv_lang).lower() + '/' ) #member url with language
                #print(office_indv_lang)
                indv_office_url_lang = base_url + indv_office_url.replace("/fr/", str(office_indv_lang).lower() + '/' ) #office url with language
                #print(indv_office_url_lang)
                indv_html = requests.get(indv_mem_url_lang) #request by mem_url
                office_indv_html = requests.get(indv_office_url_lang) #request by mem_url

                indv_doc = lxml.html.fromstring(indv_html.content)
                office_indv_doc = lxml.html.fromstring(office_indv_html.content)

                indv_full_address = indv_doc.xpath('//table//tr[2]/td/text()') #get full address block 
                office_indv_full_address = office_indv_doc.xpath('//table//tr[2]/td/text()') #get full address block 

                indv_full_address_clean = clean_list(indv_full_address)
                office_indv_full_address_clean = clean_list(office_indv_full_address)

                while(len(indv_full_address_clean)<5):
                        indv_full_address_clean.append('')
                join_indv_full_address_clean = " ".join(indv_full_address_clean)

                while(len(office_indv_full_address_clean)<4):
                        office_indv_full_address_clean.append('')
                join_office_indv_full_address_clean = " ".join(office_indv_full_address_clean)

                office_sector_list = office_indv_doc.xpath('//table//tr[6]//td//text()')
                print(office_sector_list)
                if office_sector_list:
                    office_sector = ','.join(office_sector_list)
                else:
                    office_sector = ''
                print(office_sector)

                #***implement later
                contact = indv_doc.xpath('//table//tr[4]/td[2]/div/text()')

                job = indv_doc.xpath('//table//tr[6]/td[2]/text()')
                if job:
                    job = clean_list(job)
                else:
                    job=['']
                sector = indv_doc.xpath('//table//tr[7]/td[2]/text()')
                if sector:
                    sector = clean_list(sector)
                else:
                    sector=['']
                
                group = indv_doc.xpath('//table//tr[8]/td[2]/text()')
                if group:
                    group = clean_list(group)
                else:
                    group=['']
                
                section = indv_doc.xpath('//table//tr[9]/td[2]/text()')
                if section:
                    section = clean_list(section)
                else:
                    section=['']

                #print(join_indv_full_address_clean, contact)
                #print(job, sector, group, section)
                wdf = pd.DataFrame([[ids+1, indv_mem_url_lang, indv_lang, join_indv_full_address_clean,indv_full_address_clean[0], indv_full_address_clean[1], indv_full_address_clean[2], indv_full_address_clean[3], indv_full_address_clean[4], indv_zip, contact[0], job[0], sector[0], group[0], section[0]]], columns=["ID","URL", "LANGUGE", "FULL_ADDRESS", "GENDER", "NAME", "EDUCATION", "ADDRESS", "CITY", "ZIP_CODE", "CONTACT", "JOB", "SECTOR", "GROUP", "SECTION"])

                now = datetime.now()
                dts = now.strftime("%d/%m/%Y %H:%M")
                mo_wdf = pd.DataFrame([[ids+1, ids+1, ids+1, dts]], columns=["ID","MEMBER_ID", "OFFICE_ID", "COLLECTED_AT"])

                office_wdf = pd.DataFrame([[ids+1, indv_office_url_lang, office_indv_lang, join_office_indv_full_address_clean,office_indv_full_address_clean[0], office_indv_full_address_clean[1], office_indv_full_address_clean[2], office_indv_zip, '','','','', office_sector]], columns=["ID","URL", "LANGUGE", "FULL_ADDRESS", "NAME", "ADDRESS", "CITY", "ZIP_CODE", "EMAIL", "TEL", "FAX", "WEBSITE", "SECTOR"])
                #with pd.ExcelWriter("member.xlsx") as writer:
                    #wdf.to_excel(writer, sheet_name='member', index=False)
                #print(wdf)

                if fe_flag==1:
                    wb = load_workbook(filename = "member.xlsx")
                    ws = wb["member"]
                    mows = wb["member_office"]
                    ows= wb["office"]
                    for r in dataframe_to_rows(wdf, index=False, header=False):
                        ws.append(r)
                    for d in dataframe_to_rows(mo_wdf, index=False, header=False):
                        mows.append(d)
                    for s in dataframe_to_rows(office_wdf, index=False, header=False):
                        ows.append(s)
                    wb.save("member.xlsx")
                    wb.close
                    print("page: " + str(page+1) +"  member: " + str(n+1) +"  office: " + str(n+1) +" in excel")
                else:
                    with pd.ExcelWriter("member.xlsx") as writer:
                        wdf.to_excel(writer, sheet_name='member', index=False)
                        mo_wdf.to_excel(writer, sheet_name='member_office', index=False)
                        office_wdf.to_excel(writer, sheet_name='office', index=False)
                    print('Creating New Excel')
                    fe_flag=1
                    print("page: " + str(page+1) +"  member: " + str(n+1) +"  office: " + str(n+1) +" in excel")


            else: #if zip not found
                indv_lang = 'FR'
                office_indv_lang = 'FR'
                indv_mem_url = doc.xpath('//table//a/@href')[n]
                indv_office_url = doc.xpath('//table//td[1]//a/@href')[n]

                indv_mem_url_lang = base_url + indv_mem_url.replace("/fr/", str(indv_lang).lower() + '/' ) #member url with language
                #print(office_indv_lang)
                indv_office_url_lang = base_url + indv_office_url.replace("/fr/", str(office_indv_lang).lower() + '/' ) #office url with language
                #print(indv_office_url_lang)
                indv_html = requests.get(indv_mem_url_lang) #request by mem_url
                office_indv_html = requests.get(indv_office_url_lang) #request by mem_url

                indv_doc = lxml.html.fromstring(indv_html.content)
                office_indv_doc = lxml.html.fromstring(office_indv_html.content)

                indv_full_address = indv_doc.xpath('//table//tr[2]/td/text()') #get full address block 
                office_indv_full_address = office_indv_doc.xpath('//table//tr[2]/td/text()') #get full address block 

                indv_full_address_clean = clean_list(indv_full_address)
                office_indv_full_address_clean = clean_list(office_indv_full_address)

                while(len(indv_full_address_clean)<5):
                        indv_full_address_clean.append('')
                join_indv_full_address_clean = " ".join(indv_full_address_clean)

                while(len(office_indv_full_address_clean)<4):
                        office_indv_full_address_clean.append('')
                join_office_indv_full_address_clean = " ".join(office_indv_full_address_clean)

                office_sector_list = office_indv_doc.xpath('//table//tr[6]//td//text()')
                print(office_sector_list)
                if office_sector_list:
                    office_sector = ','.join(office_sector_list)
                else:
                    office_sector = ''
                print(office_sector)

                #***implement later
                contact = indv_doc.xpath('//table//tr[4]/td[2]/div/text()')

                job = indv_doc.xpath('//table//tr[6]/td[2]/text()')
                if job:
                    job = clean_list(job)
                else:
                    job=['']
                sector = indv_doc.xpath('//table//tr[7]/td[2]/text()')
                if sector:
                    sector = clean_list(sector)
                else:
                    sector=['']
                
                group = indv_doc.xpath('//table//tr[8]/td[2]/text()')
                if group:
                    group = clean_list(group)
                else:
                    group=['']
                
                section = indv_doc.xpath('//table//tr[9]/td[2]/text()')
                if section:
                    section = clean_list(section)
                else:
                    section=['']

                #print(join_indv_full_address_clean, contact)
                #print(job, sector, group, section)
                wdf = pd.DataFrame([[ids+1, indv_mem_url_lang, indv_lang, join_indv_full_address_clean,indv_full_address_clean[0], indv_full_address_clean[1], indv_full_address_clean[2], indv_full_address_clean[3], indv_full_address_clean[4], indv_zip, contact[0], job[0], sector[0], group[0], section[0]]], columns=["ID","URL", "LANGUGE", "FULL_ADDRESS", "GENDER", "NAME", "EDUCATION", "ADDRESS", "CITY", "ZIP_CODE", "CONTACT", "JOB", "SECTOR", "GROUP", "SECTION"])

                now = datetime.now()
                dts = now.strftime("%d/%m/%Y %H:%M")
                mo_wdf = pd.DataFrame([[ids+1, ids+1, ids+1, dts]], columns=["ID","MEMBER_ID", "OFFICE_ID", "COLLECTED_AT"])

                office_wdf = pd.DataFrame([[ids+1, indv_office_url_lang, office_indv_lang, join_office_indv_full_address_clean,office_indv_full_address_clean[0], office_indv_full_address_clean[1], office_indv_full_address_clean[2], office_indv_zip, '','','','', office_sector]], columns=["ID","URL", "LANGUGE", "FULL_ADDRESS", "NAME", "ADDRESS", "CITY", "ZIP_CODE", "EMAIL", "TEL", "FAX", "WEBSITE", "SECTOR"])
                #with pd.ExcelWriter("member.xlsx") as writer:
                    #wdf.to_excel(writer, sheet_name='member', index=False)
                #print(wdf)

                if fe_flag==1:
                    wb = load_workbook(filename = "member.xlsx")
                    ws = wb["member"]
                    mows = wb["member_office"]
                    ows= wb["office"]
                    for r in dataframe_to_rows(wdf, index=False, header=False):
                        ws.append(r)
                    for d in dataframe_to_rows(mo_wdf, index=False, header=False):
                        mows.append(d)
                    for s in dataframe_to_rows(office_wdf, index=False, header=False):
                        ows.append(s)
                    wb.save("member.xlsx")
                    wb.close
                    print("page: " + str(page+1) +"  member: " + str(n+1) +"  office: " + str(n+1) +" in excel")
                else:
                    with pd.ExcelWriter("member.xlsx") as writer:
                        wdf.to_excel(writer, sheet_name='member', index=False)
                        mo_wdf.to_excel(writer, sheet_name='member_office', index=False)
                        office_wdf.to_excel(writer, sheet_name='office', index=False)
                    print('Creating New Excel')
                    fe_flag=1
                    print("page: " + str(page+1) +"  member: " + str(n+1) +"  office: " + str(n+1) +" in excel")
            ids+=1
            n+=1
        page=page+1
        n=0
def clean_list(list2clean):
    clean_list = []
    #clean full address
    for element in list2clean:
        clean_list.append(element.strip())
    return list(filter(lambda e: e != '', clean_list))

print(get_indv(page,df,fe_flag,n))